from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from .models import House, Booking, UtilityExpense,MonthlyEarning, YearlyEarning, HouseEarning, MonthlyExpense
from datetime import datetime, timedelta
import json
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
import re
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import House
from django.contrib.auth import get_user_model
from django.contrib.auth import authenticate, login as auth_login
from .forms import BookingForm
#def landing_page(request):
    #return render(request, 'landing.html')  # This will load landing.html

from django.shortcuts import render, redirect, get_object_or_404
from .models import House

User = get_user_model()

def register(request):
    if request.method == 'POST':
        username = request.POST['username']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        email = request.POST['email']
        password = request.POST['password']

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists')
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            messages.success(request, 'Account created successfully')
            return redirect('login')

    return render(request, 'houses/register.html')

from django.contrib.auth import authenticate, login as auth_login, get_user_model
from django.contrib import messages
from django.shortcuts import render, redirect

from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.shortcuts import render, redirect
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth import get_user_model

User = get_user_model()
def login_view(request):
    if request.method == 'POST':
        login_input = request.POST.get('login')
        password = request.POST.get('password')

        print(f"[DEBUG] login_input = {repr(login_input)}")

        user = User.objects.filter(email=login_input).first()
        if not user:
            user = User.objects.filter(username=login_input).first()

        print(f"[DEBUG] user = {user}")

        if user is None:
            messages.error(request, 'Email-ul sau username-ul introdus nu este înregistrat.')
            return render(request, 'houses/login.html')

        if not user.check_password(password):
            messages.error(request, 'Parola introdusă este greșită.')
            return render(request, 'houses/login.html')

        if not user.is_active:
            messages.error(request, 'Contul este dezactivat.')
            return render(request, 'houses/login.html')

        auth_login(request, user)
        return redirect('/')

    return render(request, 'houses/login.html')

# List all houses
def houses(request):
    myhouses = House.objects.all()
    return render(request, 'houses/all_houses.html', {'myhouses': myhouses})

# View to manage houses (view all houses)
def manage_houses(request):
    houses = House.objects.all()
    return render(request, 'houses/manage_houses.html', {'houses': houses})

# View to add or edit a house
def house_form(request, house_id=None):  # Use 'house_id' instead of 'id'
    if house_id:
        house = get_object_or_404(House, id=house_id)  # Get existing house for editing
    else:
        house = None  # Create a new house if no ID

    context = {'house': house}

    if request.method == 'POST':
        name = request.POST['name']
        address = request.POST['address']
        price = request.POST['price']
        photo = request.FILES.get('photo')  # Get the uploaded photo

        if house:
            # Update the existing house
            house.name = name
            house.address = address
            house.price = price
            if photo:  # If a new photo is uploaded
                house.photo = photo
            house.save()
        else:
            # Create a new house with a photo
            House.objects.create(name=name, address=address, price=price, photo=photo)

        return redirect('manage_houses')  # Redirect to the manage houses page

    return render(request, 'houses/house_form.html', context)

# View to delete a house
def delete_house(request, house_id):  # Use 'house_id' here as well
    house = get_object_or_404(House, id=house_id)  # Get the house to delete
    house.delete()
    return redirect('manage_houses')  # Redirect after deletion


# View for house details
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
from .models import House, Booking, Discount
from .utils import get_discounted_price  # Import the discount function

from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from datetime import datetime, timedelta
from .models import House, Booking
import json

@csrf_exempt  # Temporarily disable CSRF for debugging, use only during development
def house_detail(request, house_id):
    myhouse = get_object_or_404(House, id=house_id)

    if request.method == "POST":
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        customer_name = request.POST.get('customer_name', 'Anonymous')

        if not start_date_str or not end_date_str:
            return JsonResponse({"success": False, "message": "Please select start and end dates."})

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        if start_date >= end_date:
            return JsonResponse({"success": False, "message": "End date must be after start date."})

        # Overlapping booking check
        if Booking.objects.filter(house=myhouse, start_date__lt=end_date, end_date__gt=start_date).exists():
            return JsonResponse({"success": False, "message": "The selected dates are already booked."})

        # Create the booking (price will be calculated in the save method)
        try:
            new_booking = Booking.objects.create(
                house=myhouse,
                customer_name=customer_name,
                start_date=start_date,
                end_date=end_date,
                user=request.user
            )

            #new_booking.save()  # Save the booking, triggering any calculations

            return JsonResponse({"success": True, "message": "Booking successful!", "customer_name": new_booking.customer_name})
        
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Something went wrong: {str(e)}"})

    # Get booked dates
    booked_ranges = Booking.objects.filter(house=myhouse)
    booked_dates = []

    for booking in booked_ranges:
        current_date = booking.start_date
        while current_date <= booking.end_date:
            booked_dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)

    # Return data to render the page
    return render(request, 'houses/details.html', {'myhouse': myhouse, 'booked_dates': json.dumps(booked_dates)})


# Function to validate decimal input
def isValidDecimal(value):
    decimalPattern = r'^\d+(\.\d{1,2})?$'  # Pattern for valid decimal (up to two decimal places)
    return bool(re.match(decimalPattern, value))

from django.http import JsonResponse
from decimal import Decimal
from .models import UtilityExpense

import datetime
from decimal import Decimal
from django.http import JsonResponse
from .models import UtilityExpense, House
from datetime import datetime  # Ensure this is at the top of the file

from datetime import datetime
from decimal import Decimal
from django.http import JsonResponse
from django.shortcuts import render

from datetime import date  # Add this import at the top

# Your other views and logic
def add_utility_expenses(request):
    houses = House.objects.all()
    current_year = datetime.now().year  # Get the current year dynamically
    months = range(1, 13)
    years = range(current_year, current_year + 10)  # Start from current year

    if request.method == "POST":
        house_id = request.POST.get('house')
        month = request.POST.get('month')
        year = request.POST.get('year')
        water_expense = request.POST.get('water')
        electricity_expense = request.POST.get('electricity')
        proceed = request.POST.get('proceed')  # Check if user confirmed update

        if not water_expense and not electricity_expense:
            return JsonResponse({"message": "Please enter at least one expense (water or electricity)."}, status=400)
    
        # Validate input fields (only if they are not empty)
        if water_expense and not isValidDecimal(water_expense):
            return JsonResponse({"message": "Invalid water expense value."}, status=400)
        if electricity_expense and not isValidDecimal(electricity_expense):
            return JsonResponse({"message": "Invalid electricity expense value."}, status=400)

        # Convert only non-empty values to Decimal
        water_expense = Decimal(water_expense) if water_expense else None
        electricity_expense = Decimal(electricity_expense) if electricity_expense else None

        # Get the first day of the month as a DateField
        first_day_of_month = date(int(year), int(month), 1)

        # Check if an entry already exists for this house and month
        existing_entry = UtilityExpense.objects.filter(house_id=house_id, date=first_day_of_month).first()

        if existing_entry:
            # Check if the user confirmed update in the modal (proceed == "true")
            if proceed == "true":  # User confirmed update
                # Only update the water expense if provided and not None
                if water_expense is not None:
                    existing_entry.water_expense = water_expense
                
                # Only update the electricity expense if provided and not None
                if electricity_expense is not None:
                    existing_entry.electricity_expense = electricity_expense

                # Recalculate total expense
                existing_entry.total_expense = float(existing_entry.water_expense) + float(existing_entry.electricity_expense)
                existing_entry.save()

                return JsonResponse({"message": "Expenses updated successfully."})
            else:
                modal_message = ""
                
                # Only show the modal if water is being updated and it's already > 0
                if water_expense is not None and existing_entry.water_expense > 0:
                    modal_message += f"Water expense is already €{existing_entry.water_expense}. Do you wish to update it? "

                # Only show the modal if electricity is being updated and it's already > 0
                if electricity_expense is not None and existing_entry.electricity_expense > 0:
                    modal_message += f"Electricity expense is already €{existing_entry.electricity_expense}. Do you wish to update it? "
                
                # If no modal message, proceed with the update automatically (save without modal)
                if not modal_message:
                    # Update the expenses without needing confirmation
                    if water_expense is not None:
                        existing_entry.water_expense = water_expense
                    if electricity_expense is not None:
                        existing_entry.electricity_expense = electricity_expense

                    # Recalculate total expense
                    existing_entry.total_expense = float(existing_entry.water_expense) + float(existing_entry.electricity_expense)
                    existing_entry.save()
                    update_monthly_expense(house_id, first_day_of_month)  # Call the update function
                    return JsonResponse({"message": "Expenses updated successfully."})

                # If there is a modal message, return the modal message
                return JsonResponse({
                    "message": "Existing expenses found",
                    "modal_message": modal_message,
                    "existing_water": str(existing_entry.water_expense),
                    "existing_electricity": str(existing_entry.electricity_expense),
                })

        else:
            # If no existing entry for this house, month, and year, create a new entry
            UtilityExpense.objects.create(
                house_id=house_id,
                date=first_day_of_month,  # Use the full date here
                water_expense=water_expense or 0,  # Default to 0 if still None
                electricity_expense=electricity_expense or 0,
                total_expense=float((water_expense or 0) + (electricity_expense or 0))
            )
            return JsonResponse({"message": "Expenses added successfully."})

    return render(request, 'houses/utilityexpense.html', {
        'houses': houses,
        'months': months,
        'years': years,
    })

def update_monthly_expense(house_id, month_date):  # Use a DateField directly

    from django.db.models import Sum
    from datetime import date

    # Get the first day of the month
    month_date = date(year=int(year), month=int(month), day=1)

    # Sum total utility expenses for this house & month
    total_utilities = UtilityExpense.objects.filter(
        house_id=house_id,
        date=month_date
    ).aggregate(total=Sum('total_expense'))['total'] or 0  # Default to 0 if none

    # Sum total booking expenses (e.g., cleaning fees)
    total_booking_expenses = Booking.objects.filter(
        house_id=house_id,
        start_date__year=year,
        start_date__month=month
    ).aggregate(total=Sum('cleaning_fee'))['total'] or 0  # Default to 0 if none

    # Calculate total expenses
    total_expenses = total_utilities + total_booking_expenses

    # Only create/update MonthlyExpense if there's **any** expense
    if total_expenses > 0:
        monthly_expense, created = MonthlyExpense.objects.get_or_create(
            house_id=house_id,
            date=month_date,
            defaults={"total_expense": total_expenses}
        )
        if not created:
            # Update existing entry
            monthly_expense.total_expense = total_expenses
            monthly_expense.save()




from django.shortcuts import render
from decimal import Decimal
from .models import MonthlyEarning, YearlyEarning, MonthlyExpense
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt


from django.shortcuts import render
from django.db.models import Sum
from .models import MonthlyEarning, House, MonthlyExpense, UtilityExpense, HouseEarning, BookingExpense
from django.utils import timezone



from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import render

from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import render

def financial_overview(request):
    # Get the selected time period from the URL or default to "monthly"
    time_period = request.GET.get('time_period', 'monthly')

    # Get the current month and year
    current_month = datetime.now().month
    current_year = datetime.now().year

    # Determine the date range based on the time period
    if time_period == 'monthly':
        # Start of the current month
        start_date = datetime(current_year, current_month, 1)
        # End of the current month: Take the first day of the next month and subtract 1 day
        next_month = current_month + 1 if current_month < 12 else 1
        next_month_year = current_year if current_month < 12 else current_year + 1
        end_date = datetime(next_month_year, next_month, 1) - timedelta(days=1)
    elif time_period == 'quarterly-q1':
        start_date = datetime(current_year, 1, 1)
        end_date = datetime(current_year, 3, 31)
    elif time_period == 'quarterly-q2':
        start_date = datetime(current_year, 4, 1)
        end_date = datetime(current_year, 6, 30)
    elif time_period == 'quarterly-q3':
        start_date = datetime(current_year, 7, 1)
        end_date = datetime(current_year, 9, 30)
    elif time_period == 'quarterly-q4':
        start_date = datetime(current_year, 10, 1)
        end_date = datetime(current_year, 12, 31)
    elif time_period == 'yearly':
        start_date = datetime(current_year, 1, 1)
        end_date = datetime(current_year, 12, 31)
    else:
        # Default to the current month if the time period is not recognized
        start_date = datetime(current_year, current_month, 1)
        next_month = current_month + 1 if current_month < 12 else 1
        next_month_year = current_year if current_month < 12 else current_year + 1
        end_date = datetime(next_month_year, next_month, 1) - timedelta(days=1)

    # DEBUG: Print the date range for the selected period
    print(f"DEBUG: Time Period - {time_period}")
    print(f"DEBUG: Start Date: {start_date}")
    print(f"DEBUG: End Date: {end_date}")

    # Get total earnings for the selected time period
    total_earnings = MonthlyEarning.objects.filter(
        month_name__gte=start_date, 
        month_name__lte=end_date
    ).aggregate(Sum('total_earnings'))['total_earnings__sum'] or 0.00

    # DEBUG: Print the raw total earnings value
    print(f"DEBUG: Total Earnings from DB: {total_earnings}")

    total_earnings_decimal = Decimal(total_earnings)

    # Calculate total VAT collected (19% of total earnings)
    total_vat_collected = total_earnings_decimal * Decimal('0.19')

    # DEBUG: Print the calculated VAT collected
    print(f"DEBUG: Total VAT Collected: {total_vat_collected}")

    # Get total expenses for the selected time period
    total_expenses = MonthlyExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    # DEBUG: Print the raw total expenses value
    print(f"DEBUG: Total Expenses from DB: {total_expenses}")

    # Get total utility expenses for the selected time period
    total_utility_expenses = UtilityExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    # DEBUG: Print the raw utility expenses value
    print(f"DEBUG: Total Utility Expenses from DB: {total_utility_expenses}")

    # Get total booking expenses for the selected time period
    total_booking_expenses = BookingExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0.00

    # DEBUG: Print the raw booking expenses value
    print(f"DEBUG: Total Booking Expenses from DB: {total_booking_expenses}")

    # Combine all expenses
    total_net_expenses = Decimal(total_utility_expenses) + Decimal(total_booking_expenses)

    # DEBUG: Print the total combined expenses
    print(f"DEBUG: Total Net Expenses: {total_net_expenses}")

    # Calculate net earnings (total earnings - total expenses)
    total_net_earnings = total_earnings_decimal - total_net_expenses

    # DEBUG: Print the net earnings
    print(f"DEBUG: Total Net Earnings: {total_net_earnings}")

    # Calculate earnings per house (Excluding and Including VAT)
    houses = House.objects.all()
    house_earnings_data = []
    total_earnings_per_house = Decimal('0.00')

    for house in houses:
        house_earnings_excl_vat = HouseEarning.objects.filter(
            house=house, 
            month__gte=start_date, 
            month__lte=end_date
        ).aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')

        # DEBUG: Print the earnings for each house
        print(f"DEBUG: Earnings for House {house.name}: {house_earnings_excl_vat}")

        # Calculate earnings including VAT (19% VAT)
        house_earnings_incl_vat = house_earnings_excl_vat * Decimal('1.19')

        house_earnings_data.append({
            'house_name': house.name,
            'total_earnings_excl_vat': house_earnings_excl_vat,
            'total_earnings_incl_vat': house_earnings_incl_vat
        })

        # Add to total earnings across all houses
        total_earnings_per_house += house_earnings_excl_vat

    # Calculate the average earnings per house
    # Calculate the average earnings per house
    avg_earnings_per_house = round(total_earnings_per_house / len(houses), 2) if len(houses) > 0 else 0.00


    # DEBUG: Print the total earnings per house and average earnings
    print(f"DEBUG: Total Earnings Across All Houses: {total_earnings_per_house}")
    print(f"DEBUG: Average Earnings Per House: {avg_earnings_per_house}")

    # Get total VAT deductible (e.g., assume 19% for simplicity on total expenses)
    total_vat_deductible = total_net_expenses * Decimal('0.19')

    # DEBUG: Print the VAT deductible
    print(f"DEBUG: Total VAT Deductible: {total_vat_deductible}")

    # Calculate net VAT
    net_vat = total_vat_collected - total_vat_deductible

    # DEBUG: Print the net VAT
    print(f"DEBUG: Net VAT: {net_vat}")

    # Pass all data to the template
    return render(request, 'houses/financial_overview.html', {
        'total_earnings': total_earnings_decimal,
        'total_expenses': total_expenses,
        'total_utility_expenses': total_utility_expenses,
        'total_booking_expenses': total_booking_expenses,
        'total_net_earnings': total_net_earnings,
        'total_vat_collected': total_vat_collected,
        'total_vat_deductible': total_vat_deductible,
        'houses': houses,
        'total_net_expenses': total_net_expenses,
        'net_vat': net_vat,
        'house_earnings_data': house_earnings_data,
        'avg_earnings_per_house': avg_earnings_per_house,  # Make sure this is passed
        'time_period': time_period,  # Pass selected time period to the template
    })


@csrf_exempt
def calculate_taxes(request):
    # Get the latest monthly and yearly earnings or set defaults if no entries
    try:
        total_monthly_earnings = MonthlyEarning.objects.latest('month_name').total_earnings
    except MonthlyEarning.DoesNotExist:
        total_monthly_earnings = Decimal('0.00')

    try:
        total_yearly_earnings = YearlyEarning.objects.latest('year').total_earnings
    except YearlyEarning.DoesNotExist:
        total_yearly_earnings = Decimal('0.00')

    # Fetch total yearly expenses for PFA (from MonthlyExpense)
    try:
        total_yearly_expenses = MonthlyExpense.objects.aggregate(total=Sum('total_expense'))['total'] or Decimal('0.00')
    except MonthlyExpense.DoesNotExist:
        total_yearly_expenses = Decimal('0.00')

    if request.method == 'POST':
        data = json.loads(request.body)
        business_mode = data.get('business_mode')
        num_employees = int(data.get('employees', 0))

        # Initialize response data dictionary
        response_data = {}

        if business_mode == 'PFA':
            # Calculate net profit (earnings - expenses)
            total_profit = total_yearly_earnings - total_yearly_expenses

            # Apply income tax (10% on profit) for PFA
            income_tax = total_profit * Decimal('0.1')  # 10% income tax

            # Apply VAT (9% on profit) for PFA
            vat = total_profit * Decimal('0.09')  # 9% VAT on profit

            # Prepare response data for PFA
            response_data = {
                'income_tax': income_tax.quantize(Decimal('0.01')),
                'vat': vat.quantize(Decimal('0.01'))
            }

        elif business_mode == 'SRL':
            # Use total monthly earnings for SRL
            total_income = total_monthly_earnings

            # Determine microenterprise tax rate based on number of employees (1% or 3%)
            if num_employees == 0:
                micro_tax = total_income * Decimal('0.01')  # 1% tax rate for SRL without employees
            else:
                micro_tax = total_income * Decimal('0.03')  # 3% tax rate for SRL with employees

            # Apply VAT (9% on monthly earnings) for SRL
            vat = total_income * Decimal('0.09')  # 9% VAT for SRL

            # Prepare response data for SRL
            response_data = {
                'micro_tax': micro_tax.quantize(Decimal('0.01')),
                'vat': vat.quantize(Decimal('0.01'))
            }

        # Return the JSON response with calculated taxes
        return JsonResponse(response_data)

import csv
from datetime import datetime, timedelta
from decimal import Decimal
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
from .models import MonthlyEarning, MonthlyExpense, UtilityExpense, BookingExpense, House, HouseEarning

def export_to_csv(request):
    time_period = request.GET.get('time_period', 'monthly')
    current_month = datetime.now().month
    current_year = datetime.now().year

    if time_period == 'monthly':
        start_date = datetime(current_year, current_month, 1)
        next_month = current_month + 1 if current_month < 12 else 1
        next_month_year = current_year if current_month < 12 else current_year + 1
        end_date = datetime(next_month_year, next_month, 1) - timedelta(days=1)
    elif time_period == 'quarterly-q1':
        start_date = datetime(current_year, 1, 1)
        end_date = datetime(current_year, 3, 31)
    elif time_period == 'quarterly-q2':
        start_date = datetime(current_year, 4, 1)
        end_date = datetime(current_year, 6, 30)
    elif time_period == 'quarterly-q3':
        start_date = datetime(current_year, 7, 1)
        end_date = datetime(current_year, 9, 30)
    elif time_period == 'quarterly-q4':
        start_date = datetime(current_year, 10, 1)
        end_date = datetime(current_year, 12, 31)
    elif time_period == 'yearly':
        start_date = datetime(current_year, 1, 1)
        end_date = datetime(current_year, 12, 31)
    else:
        start_date = datetime(current_year, current_month, 1)
        next_month = current_month + 1 if current_month < 12 else 1
        next_month_year = current_year if current_month < 12 else current_year + 1
        end_date = datetime(next_month_year, next_month, 1) - timedelta(days=1)

    total_earnings = MonthlyEarning.objects.filter(
        month_name__gte=start_date, 
        month_name__lte=end_date
    ).aggregate(Sum('total_earnings'))['total_earnings__sum'] or 0.00

    total_earnings_decimal = Decimal(total_earnings)
    total_vat_collected = total_earnings_decimal * Decimal('0.19')

    total_expenses = MonthlyExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    total_utility_expenses = UtilityExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    total_booking_expenses = BookingExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0.00

    total_net_expenses = Decimal(total_utility_expenses) + Decimal(total_booking_expenses)
    total_net_earnings = total_earnings_decimal - total_net_expenses

    houses = House.objects.all()
    house_earnings_data = []

    for house in houses:
        house_earnings_excl_vat = HouseEarning.objects.filter(
            house=house, 
            month__year=current_year, 
            month__month=current_month
        ).aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')

        house_earnings_incl_vat = house_earnings_excl_vat * Decimal('1.19')
        house_earnings_data.append([house.name, house_earnings_excl_vat, house_earnings_incl_vat])

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="financial_overview_{time_period}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Total Earnings', 'Total Expenses', 'Total Utility Expenses', 'Total Booking Expenses', 'Total Net Earnings', 'Total VAT Collected'])
    writer.writerow([total_earnings_decimal, total_expenses, total_utility_expenses, total_booking_expenses, total_net_earnings, total_vat_collected])

    writer.writerow([])
    writer.writerow(['House Name', 'Earnings Excluding VAT', 'Earnings Including VAT'])
    for row in house_earnings_data:
        writer.writerow(row)

    return response


import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum
from .models import MonthlyExpense, MonthlyEarning, UtilityExpense, BookingExpense, House, HouseEarning

def export_to_excel(request):
    # Get the selected time period from the request
    time_period = request.GET.get('time_period', 'monthly')

    # Get the current month and year
    current_month = datetime.now().month
    current_year = datetime.now().year

    # Determine the date range
    if time_period == 'monthly':
        start_date = datetime(current_year, current_month, 1)
        next_month = current_month + 1 if current_month < 12 else 1
        next_month_year = current_year if current_month < 12 else current_year + 1
        end_date = datetime(next_month_year, next_month, 1) - timedelta(days=1)
    elif time_period == 'quarterly-q1':
        start_date, end_date = datetime(current_year, 1, 1), datetime(current_year, 3, 31)
    elif time_period == 'quarterly-q2':
        start_date, end_date = datetime(current_year, 4, 1), datetime(current_year, 6, 30)
    elif time_period == 'quarterly-q3':
        start_date, end_date = datetime(current_year, 7, 1), datetime(current_year, 9, 30)
    elif time_period == 'quarterly-q4':
        start_date, end_date = datetime(current_year, 10, 1), datetime(current_year, 12, 31)
    elif time_period == 'yearly':
        start_date, end_date = datetime(current_year, 1, 1), datetime(current_year, 12, 31)

    # Fetch data
    total_earnings = MonthlyEarning.objects.filter(
        month_name__gte=start_date, 
        month_name__lte=end_date
    ).aggregate(Sum('total_earnings'))['total_earnings__sum'] or 0.00

    total_expenses = MonthlyExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    total_utility_expenses = UtilityExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    total_booking_expenses = BookingExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0.00

    total_net_expenses = Decimal(total_utility_expenses) + Decimal(total_booking_expenses)
    total_earnings_decimal = Decimal(total_earnings)
    total_net_earnings = total_earnings_decimal - total_net_expenses

    total_vat_collected = total_earnings_decimal * Decimal('0.19')
    total_vat_deductible = total_net_expenses * Decimal('0.19')
    net_vat = total_vat_collected - total_vat_deductible

    # Earnings per house
    houses = House.objects.all()
    house_earnings_data = []
    total_earnings_per_house = Decimal('0.00')

    for house in houses:
        house_earnings_excl_vat = HouseEarning.objects.filter(
            house=house, 
            month__year=current_year, 
            month__month=current_month
        ).aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')

        house_earnings_incl_vat = house_earnings_excl_vat * Decimal('1.19')

        house_earnings_data.append({
            'house_name': house.name,
            'total_earnings_excl_vat': house_earnings_excl_vat,
            'total_earnings_incl_vat': house_earnings_incl_vat
        })

        total_earnings_per_house += house_earnings_excl_vat

    # Calculate the average earnings per house
    avg_earnings_per_house = round(total_earnings_per_house / len(houses), 2) if len(houses) > 0 else 0.00


    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Overview"

    bold_font = Font(bold=True)

    # 🟢 Financial Summary
    ws.append(["Financial Overview"])
    ws["A1"].font = bold_font
    ws.append([])
    
    summary_data = [
        ("Total Earnings", total_earnings_decimal),
        ("Total Expenses", total_net_expenses),
        ("Net Earnings", total_net_earnings),
        ("Total VAT Collected", total_vat_collected),
        ("Total VAT Deductible", total_vat_deductible),
        ("Net VAT", net_vat),
        ("Average Earnings per House", avg_earnings_per_house),
    ]
    
    for row in summary_data:
        ws.append(row)

    ws.append([])

    # 🟢 Expense Breakdown
    ws.append(["Expense Breakdown"])
    ws.append(["Expense Type", "Amount"])
    ws["A{}".format(ws.max_row)].font = bold_font

    ws.append(["Utility Expenses", total_utility_expenses])
    ws.append(["Booking Expenses", total_booking_expenses])
    ws.append(["Total Expenses", total_net_expenses])

    ws.append([])

    # 🟢 Earnings by House
    ws.append(["Earnings by House"])
    ws.append(["House Name", "Total Earnings (Excl. VAT)", "Total Earnings (Incl. VAT)"])
    ws["A{}".format(ws.max_row)].font = bold_font

    for house_data in house_earnings_data:
        ws.append([
            house_data['house_name'],
            house_data['total_earnings_excl_vat'],
            house_data['total_earnings_incl_vat']
        ])

    ws.append([])

    # 🟢 VAT Overview
    ws.append(["VAT Overview"])
    ws.append(["VAT Type", "Value"])
    ws["A{}".format(ws.max_row)].font = bold_font

    vat_data = [
        ("Total VAT Collected", total_vat_collected),
        ("Total VAT Deductible", total_vat_deductible),
        ("Net VAT", net_vat),
    ]

    for row in vat_data:
        ws.append(row)

    # 📤 Save and return response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="financial_report.xlsx"'
    wb.save(response)

    return response


# Generate PDF Report
from django.shortcuts import render
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
from datetime import datetime, timedelta
from decimal import Decimal
from .models import MonthlyExpense, MonthlyEarning, UtilityExpense, BookingExpense, House, HouseEarning

def generate_pdf_report(request):
    # Get the selected time period from the URL or default to "monthly"
    time_period = request.GET.get('time_period', 'monthly')

    # Get the current month and year
    current_month = datetime.now().month
    current_year = datetime.now().year

    # Determine the date range based on the time period
    if time_period == 'monthly':
        start_date = datetime(current_year, current_month, 1)
        next_month = current_month + 1 if current_month < 12 else 1
        next_month_year = current_year if current_month < 12 else current_year + 1
        end_date = datetime(next_month_year, next_month, 1) - timedelta(days=1)
    elif time_period == 'quarterly-q1':
        start_date = datetime(current_year, 1, 1)
        end_date = datetime(current_year, 3, 31)
    elif time_period == 'quarterly-q2':
        start_date = datetime(current_year, 4, 1)
        end_date = datetime(current_year, 6, 30)
    elif time_period == 'quarterly-q3':
        start_date = datetime(current_year, 7, 1)
        end_date = datetime(current_year, 9, 30)
    elif time_period == 'quarterly-q4':
        start_date = datetime(current_year, 10, 1)
        end_date = datetime(current_year, 12, 31)
    elif time_period == 'yearly':
        start_date = datetime(current_year, 1, 1)
        end_date = datetime(current_year, 12, 31)
    else:
        start_date = datetime(current_year, current_month, 1)
        next_month = current_month + 1 if current_month < 12 else 1
        next_month_year = current_year if current_month < 12 else current_year + 1
        end_date = datetime(next_month_year, next_month, 1) - timedelta(days=1)

    # Calculate total earnings for the selected time period
    total_earnings = MonthlyEarning.objects.filter(
        month_name__gte=start_date, 
        month_name__lte=end_date
    ).aggregate(Sum('total_earnings'))['total_earnings__sum'] or 0.00
    total_earnings_decimal = Decimal(total_earnings)

    # Calculate total VAT collected (19% of total earnings)
    total_vat_collected = total_earnings_decimal * Decimal('0.19')

    # Get total expenses for the selected time period
    total_expenses = MonthlyExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    # Get total utility expenses for the selected time period
    total_utility_expenses = UtilityExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    # Get total booking expenses for the selected time period
    total_booking_expenses = BookingExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0.00

    # Combine all expenses
    total_net_expenses = Decimal(total_utility_expenses) + Decimal(total_booking_expenses)

    # Calculate net earnings (total earnings - total expenses)
    total_net_earnings = total_earnings_decimal - total_net_expenses

    # Calculate earnings per house (Excluding and Including VAT)
    houses = House.objects.all()
    house_earnings_data = []
    total_earnings_per_house = Decimal('0.00')

    for house in houses:
        house_earnings_excl_vat = HouseEarning.objects.filter(
            house=house, 
            month__year=current_year, 
            month__month=current_month
        ).aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')

        # Calculate earnings including VAT (19% VAT)
        house_earnings_incl_vat = house_earnings_excl_vat * Decimal('1.19')

        house_earnings_data.append({
            'house_name': house.name,
            'total_earnings_excl_vat': house_earnings_excl_vat,
            'total_earnings_incl_vat': house_earnings_incl_vat
        })

        # Add to total earnings across all houses
        total_earnings_per_house += house_earnings_excl_vat

    # Calculate the average earnings per house
    # Calculate the average earnings per house
    avg_earnings_per_house = round(total_earnings_per_house / len(houses), 2) if len(houses) > 0 else 0.00


    # Get total VAT deductible (e.g., assume 19% for simplicity on total expenses)
    total_vat_deductible = total_net_expenses * Decimal('0.19')

    # Calculate net VAT
    net_vat = total_vat_collected - total_vat_deductible

    # Prepare the context for rendering the PDF
    context = {
        'total_earnings': total_earnings_decimal,
        'total_expenses': total_expenses,
        'total_utility_expenses': total_utility_expenses,
        'total_booking_expenses': total_booking_expenses,
        'total_net_earnings': total_net_earnings,
        'total_vat_collected': total_vat_collected,
        'total_vat_deductible': total_vat_deductible,
        'houses': houses,
        'total_net_expenses': total_net_expenses,
        'net_vat': net_vat,
        'house_earnings_data': house_earnings_data,
        'avg_earnings_per_house': avg_earnings_per_house,
        'time_period': time_period,
    }

    # Render HTML template for the PDF
    html_string = render_to_string('pdf_report_template.html', context)

    # Generate the PDF from HTML string
    html = HTML(string=html_string)
    pdf = html.write_pdf()

    # Return PDF as response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="financial_overview_report.pdf"'

    return response

from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from .models import Booking
from decimal import Decimal
from django.core.exceptions import PermissionDenied

def booking_list(request):
    bookings = Booking.objects.all()  # Fetch all bookings
    
    return render(request, 'houses/booking_list.html', {'bookings': bookings})

def edit_booking(request, booking_id):
    if not request.user.is_staff and not request.user.is_superuser:
        raise PermissionDenied("You do not have permission to edit bookings.")
    booking = get_object_or_404(Booking, id=booking_id)
    if request.method == 'POST':
        form = BookingForm(request.POST, instance=booking)
        if form.is_valid():
            form.save()
            return redirect('booking_list')
    else:
        form = BookingForm(instance=booking)
    return render(request, 'houses/edit_booking.html', {'form': form, 'booking': booking})

def delete_booking(request, booking_id):
    if not request.user.is_staff and not request.user.is_superuser:
        raise PermissionDenied("You do not have permission to delete bookings.")
    booking = get_object_or_404(Booking, id=booking_id)
    if request.method == 'POST':
        booking.delete()
        return redirect('booking_list')
    return render(request, 'houses/delete_booking.html', {'booking': booking})

def add_note(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    if request.method == 'POST':
        note = request.POST.get('note')
        booking.note = note
        booking.save()
        return redirect('booking_list')
    return render(request, 'houses/add_note.html', {'booking': booking})


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Booking
from decimal import Decimal
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from datetime import datetime

def generate_invoice(request, booking_id):
    # Fetch the booking based on the provided booking ID
    booking = get_object_or_404(Booking, id=booking_id)
    
    # Get the booking price (excluding VAT)
    price_excl_vat = booking.booking_earnings

    # VAT rate (19%)
    vat_rate = Decimal('0.19')

    # Calculate price including VAT
    price_incl_vat = price_excl_vat * (Decimal('1') + vat_rate)

    # Calculate the number of nights (including the start day)
    total_days = (booking.end_date - booking.start_date).days + 1
    
    # Create the PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{booking.id}.pdf"'

    # Create the PDF using ReportLab
    p = canvas.Canvas(response, pagesize=letter)
    
    # Set up basic font styles
    p.setFont("Helvetica-Bold", 14)
    
    # Add Company Info (Header)
    p.drawString(100, 750, "Your Company Name")
    p.setFont("Helvetica", 10)
    p.drawString(100, 735, "Your Address Line 1")
    p.drawString(100, 720, "Your Address Line 2")
    p.drawString(100, 705, "City, Postal Code")
    p.drawString(100, 690, "Email: your@email.com | Phone: +123456789")
    
    # Add a horizontal line to separate header from the rest of the content
    p.setStrokeColor(colors.black)
    p.setLineWidth(1)
    p.line(100, 680, 500, 680)

    # Add Invoice Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 660, "Invoice")

    # Add Invoice Number and Date
    p.setFont("Helvetica", 10)
    p.drawString(400, 660, f"Invoice Number: {booking.id}")
    p.drawString(400, 645, f"Invoice Date: {datetime.now().strftime('%Y-%m-%d')}")

    # Add Booking Details
    p.setFont("Helvetica", 12)
    p.drawString(100, 620, f"Booking ID: {booking.id}")
    p.drawString(100, 600, f"Customer Name: {booking.customer_name}")
    p.drawString(100, 580, f"Start Date: {booking.start_date}")
    p.drawString(100, 560, f"End Date: {booking.end_date}")
    
    # Add Description of Service with Number of Nights
    p.drawString(100, 540, f"Description: Accommodation at {booking.house.name} for {total_days} nights ({booking.start_date} to {booking.end_date})")

    # Add Price with VAT
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, 510, f"Price (Including VAT):")
    p.setFont("Helvetica", 12)
    p.drawString(250, 510, f"{price_incl_vat:.2f}")

    # Add a footer with a thank you note or website (Optional)
    p.setFont("Helvetica", 10)
    p.drawString(100, 100, "Thank you for your business! Visit our website at www.yourcompany.com")

    # Finalize the PDF
    p.showPage()
    p.save()

    return response
from django.shortcuts import render
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Avg, F
from .models import MonthlyEarning, HouseEarning, Booking, UtilityExpense, BookingExpense, House
from django.utils import timezone
from django.db.models import ExpressionWrapper, IntegerField
from decimal import Decimal
from django.db.models import Case, When, DateField
from django.db.models.functions import Greatest, Least


def landing_page(request):
    current_month = datetime.now().month
    current_year = datetime.now().year

 # Get the full month name using strftime
    current_month_name = datetime(current_year, current_month, 1).strftime('%B')

    start_date = datetime(current_year, current_month, 1)
    next_month = current_month + 1 if current_month < 12 else 1
    next_month_year = current_year if current_month < 12 else current_year + 1
    end_date = datetime(next_month_year, next_month, 1) - timedelta(days=1)

    # Total earnings for the current month
    total_earnings = MonthlyEarning.objects.filter(
        month_name__gte=start_date, 
        month_name__lte=end_date
    ).aggregate(Sum('total_earnings'))['total_earnings__sum'] or 0.00

    # Total expenses for the current month (sum of all expenses)
    total_expenses = UtilityExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('total_expense'))['total_expense__sum'] or 0.00

    total_booking_expenses = BookingExpense.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0.00

    # Calculate net earnings (total earnings - total expenses)
    total_net_expenses = Decimal(total_expenses) + Decimal(total_booking_expenses)
    total_net_earnings = Decimal(total_earnings) - total_net_expenses

    # House earnings for the current month
    house_earnings = HouseEarning.objects.filter(
        month__gte=start_date, 
        month__lte=end_date
    ).all()

    # Upcoming bookings
    upcoming_bookings = Booking.objects.filter(start_date__gte=datetime.now()).order_by("start_date")

    # Occupancy Rate Calculation
    #total_nights_in_month = (end_date - start_date).days + 1
    #total_houses = House.objects.count()

    #booked_nights = Booking.objects.filter(
        #start_date__lte=end_date,
        #end_date__gte=start_date
    #).annotate(nights=ExpressionWrapper(F('end_date') - F('start_date') + 1, output_field=IntegerField())) \
    #.aggregate(total_booked_nights=Sum('nights'))['total_booked_nights'] or 0

    #occupancy_rate = (booked_nights / (total_nights_in_month * total_houses) * 100) if total_houses > 0 else 0
    #occupancy_rate = round(occupancy_rate, 2)  # Now it works correctly!
    # Bookings overlapping with current month
    bookings_in_month = Booking.objects.filter(
        start_date__lte=end_date,
        end_date__gte=start_date
    ).annotate(
        # Get actual nights booked within the current month
        adjusted_start=Greatest(F('start_date'), start_date),
        adjusted_end=Least(F('end_date'), end_date),
    ).annotate(
        nights=ExpressionWrapper(
            F('adjusted_end') - F('adjusted_start') + timedelta(days=1),
            output_field=IntegerField()
        )
    )

    booked_nights = bookings_in_month.aggregate(total_booked_nights=Sum('nights'))['total_booked_nights'] or 0

    # Total available nights = days in month * number of houses
    total_nights_in_month = (end_date - start_date).days + 1
    total_houses = House.objects.count()

    occupancy_rate = (booked_nights / (total_nights_in_month * total_houses) * 100) if total_houses > 0 else 0
    occupancy_rate = round(occupancy_rate, 2)
    


    # Most Frequently Booked Property for the current month
    most_booked_property = Booking.objects.filter(
        start_date__gte=start_date,
        start_date__lte=end_date
    ).values('house__name').annotate(count=Count('id')).order_by('-count').first()

    most_booked_property_name = most_booked_property['house__name'] if most_booked_property else "N/A"


    # Average Booking Duration
    #avg_booking_duration = Booking.objects.annotate(nights=(F('end_date') - F('start_date') + timedelta(days=1))) \
    #.aggregate(Avg('nights'))['nights__avg'] or 0
    # Average Booking Duration for the current month
    avg_booking_duration = Booking.objects.filter(
        start_date__gte=start_date,
        start_date__lte=end_date
    ).annotate(
        nights=ExpressionWrapper(F('end_date') - F('start_date') + timedelta(days=1), output_field=DurationField())
    ).aggregate(
        avg_nights=Avg('nights')
    )['avg_nights'] or timedelta(days=0)

    # Convert to days and round
    avg_booking_duration = avg_booking_duration.days if isinstance(avg_booking_duration, timedelta) else avg_booking_duration
    avg_booking_duration = round(avg_booking_duration, 1)


    # Ensure avg_booking_duration is a valid number
    if isinstance(avg_booking_duration, timedelta):
        avg_booking_duration = avg_booking_duration.days
    avg_booking_duration = round(avg_booking_duration, 1)

    # Recent Expenses (Fetching from UtilityExpense and BookingExpense)
    recent_utility_expenses = UtilityExpense.objects.order_by('-date')[:5]
    recent_booking_expenses = BookingExpense.objects.order_by('-date')[:5]

    # Merge and sort expenses by date (keep latest 5)
    recent_expenses = sorted(
        list(recent_utility_expenses) + list(recent_booking_expenses),
        key=lambda expense: expense.date,
        reverse=True
    )[:5]

    context = {
        'total_earnings': total_earnings,
        'total_net_earnings': total_net_earnings,  # Pass net earnings (profit) to the template
        'house_earnings': house_earnings,
        'upcoming_bookings': upcoming_bookings,
        'occupancy_rate': occupancy_rate,
        'most_booked_property': most_booked_property_name,
        'avg_booking_duration': avg_booking_duration,
        'recent_expenses': recent_expenses,
        'current_month_name': current_month_name,
    }

    return render(request, 'landing.html', context)

# views.py

from datetime import datetime
import calendar
from decimal import Decimal
from django.shortcuts import render
from .models import House, BookingExpense, UtilityExpense
from django.db.models import Sum

from django.db.models import Sum
from decimal import Decimal
from django.db.models import Sum
from decimal import Decimal
from django.db.models import Sum
from decimal import Decimal
import calendar
from datetime import datetime

def expense_overview(request):
    # Get the current date
    current_date = datetime.now()
    current_month = current_date.month
    current_year = current_date.year

    # Get house_id from the URL parameters
    house_id = request.GET.get('house', None)
    house_id = int(house_id) if house_id and house_id.isdigit() else None

    # Get selected month and year from the URL parameters
    selected_month = int(request.GET.get('month', current_month))
    selected_year = int(request.GET.get('year', current_year))
    

    # Get list of houses
    houses = House.objects.all()

    # Prepare list of months for the dropdown
    months = [{'value': i, 'name': calendar.month_name[i]} for i in range(1, 13)]

    # Prepare list of years for the dropdown (next 5 years + current year)
    years = [current_year + i for i in range(6)]

    # Function to calculate total expenses
    def calculate_total_expenses(house_id, selected_month, selected_year):
        booking_filter = {
            'date__year': selected_year,
            'date__month': selected_month
        }
        if house_id:
            booking_filter['booking__house_id'] = house_id

        booking_expenses = BookingExpense.objects.filter(**booking_filter).aggregate(total=Sum('amount'))['total'] or 0

        utility_filter = {
            'date__year': selected_year,
            'date__month': selected_month
        }
        if house_id:
            utility_filter['house_id'] = house_id

        utility_expenses = UtilityExpense.objects.filter(**utility_filter).aggregate(total=Sum('total_expense'))['total'] or 0

        return booking_expenses + utility_expenses

    # Calculate VAT deductible
    def calculate_vat_deductible(house_id, selected_month, selected_year):
        total_expenses = calculate_total_expenses(house_id, selected_month, selected_year)
        vat_rate = Decimal('0.19')
        return total_expenses * vat_rate

    # Calculate total expenses
    total_expenses = calculate_total_expenses(house_id, selected_month, selected_year)

    # Calculate utility expenses for water and electricity separately
    total_water_expenses = UtilityExpense.objects.filter(
        date__year=selected_year,
        date__month=selected_month,
        **({'house_id': house_id} if house_id else {})
    ).aggregate(total=Sum('water_expense'))['total'] or 0

    total_electricity_expenses = UtilityExpense.objects.filter(
        date__year=selected_year,
        date__month=selected_month,
        **({'house_id': house_id} if house_id else {})
    ).aggregate(total=Sum('electricity_expense'))['total'] or 0

    # Calculate total booking expenses
    total_booking_expenses = BookingExpense.objects.filter(
        date__year=selected_year,
        date__month=selected_month,
        **({'booking__house_id': house_id} if house_id else {})
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Calculate VAT deductible
    total_vat_deductible = calculate_vat_deductible(house_id, selected_month, selected_year)

    # Calculate total expenses for each house (for the graph)
    house_expenses = []
    for house in houses:
        water_expenses = UtilityExpense.objects.filter(
            house=house,
            date__year=selected_year,
            date__month=selected_month
        ).aggregate(total=Sum('water_expense'))['total'] or 0

        electricity_expenses = UtilityExpense.objects.filter(
            house=house,
            date__year=selected_year,
            date__month=selected_month
        ).aggregate(total=Sum('electricity_expense'))['total'] or 0

        booking_expenses = BookingExpense.objects.filter(
            booking__house=house,
            date__year=selected_year,
            date__month=selected_month
        ).aggregate(total=Sum('amount'))['total'] or 0

        house_total_expenses = calculate_total_expenses(house.id, selected_month, selected_year)
        house_expenses.append({
        'name': house.name,
        'total_expenses': house_total_expenses
    })



    # Pass variables to the template
    context = {
        'houses': houses,
        'house_id': house_id,
        'months': months,
        'selected_month': selected_month,
        'years': years,
        'selected_year': selected_year,
        'total_expenses': total_expenses,
        'total_utility_expenses': total_water_expenses + total_electricity_expenses,
        'total_booking_expenses': total_booking_expenses,
        'total_vat_deductible': total_vat_deductible,
        'total_water_expenses': total_water_expenses,
        'total_electricity_expenses': total_electricity_expenses,
        'house_expenses': house_expenses,  # New data for the graph
        
    }

    return render(request, 'houses/expense_overview.html', context)


from datetime import datetime, timedelta
from django.shortcuts import render
from django.db.models import Sum, Count, F, ExpressionWrapper, FloatField, Q
from decimal import Decimal
import calendar
from .models import House, Booking, MonthlyExpense

def calculate_profit(house_id=None, month=None, year=None):
    filters = {}

    if house_id:
        filters["house_id"] = house_id
    if year:
        filters["start_date__year"] = year
    if month and month != 0:
        filters["start_date__month"] = month

    total_earnings = Booking.objects.filter(**filters).aggregate(total=Sum('booking_earnings'))['total'] or Decimal('0.00')

    expense_filters = {"house_id": house_id, "date__year": year}
    if month and month != 0:
        expense_filters["date__month"] = month

    total_expenses = MonthlyExpense.objects.filter(**expense_filters).aggregate(total=Sum('total_expense'))['total'] or Decimal('0.00')

    profit = total_earnings - total_expenses

    print(f"DEBUG - Profit Calculation for house {house_id}, month {month}, year {year}:")
    print(f"Total Earnings: {total_earnings}, Total Expenses: {total_expenses}, Profit: {profit}")

    return {
        "total_earnings": total_earnings,
        "total_expenses": total_expenses,
        "profit": profit
    }

def calculate_occupancy_rate(house_id, month, year):
    if not house_id:
        return None

    # Get the start and end dates of the selected month
    first_day = datetime(year, month, 1).date()  # Convert to date
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    print(f"First Day: {first_day}, Last Day: {last_day}")  # Debugging: print the date range

    # Get all bookings for the selected house within the date range
    bookings = Booking.objects.filter(
        house_id=house_id,
        start_date__lte=last_day,
        end_date__gte=first_day
    )

    booked_days = 0
    for booking in bookings:
        # Convert booking start and end dates to date if they are datetime
        start_date = booking.start_date.date() if isinstance(booking.start_date, datetime) else booking.start_date
        end_date = booking.end_date.date() if isinstance(booking.end_date, datetime) else booking.end_date

        print(f"Booking: {start_date} to {end_date}")  # Debugging: print each booking's date range
        
        # Calculate the booked days for this booking
        booked_days += (min(end_date, last_day) - max(start_date, first_day)).days + 1
        print(f"Booked Days for this booking: {(min(end_date, last_day) - max(start_date, first_day)).days + 1}")  # Debugging

    total_days = (last_day - first_day).days + 1

    print(f"Total Days in Month: {total_days}, Booked Days: {booked_days}")  # Debugging: check total days and booked days

    occupancy_rate = round((booked_days / total_days) * 100, 2) if total_days > 0 else 0
    print(f"Occupancy Rate: {occupancy_rate}")  # Debugging: print final occupancy rate

    return occupancy_rate



from django.db.models import F, ExpressionWrapper, DurationField

def calculate_adr(house_id, month, year):
    if not house_id:
        return None

    filters = {"house_id": house_id, "start_date__year": year}
    if month and month != 0:
        filters["start_date__month"] = month

    # Get the total earnings
    total_earnings = Booking.objects.filter(**filters).aggregate(total=Sum('booking_earnings'))['total'] or Decimal('0.00')

    # Debugging: Check total earnings
    print(f"Total earnings for house {house_id} in {month}/{year}: {total_earnings}")

    # Calculate the total nights using DateDifference
    total_nights = Booking.objects.filter(**filters).annotate(
        nights=ExpressionWrapper(
            F('end_date') - F('start_date'), output_field=DurationField()
        )
    ).aggregate(total_nights=Sum('nights'))['total_nights']

    # Debugging: Check total nights before adjustment
    print(f"Total nights for house {house_id} in {month}/{year} (before adjustment): {total_nights}")

    # Adjust total_nights to include both check-in and check-out days
    total_nights = 0
    for booking in Booking.objects.filter(**filters):
        booked_days = (booking.end_date - booking.start_date).days + 1  # Add 1 to include checkout day
        total_nights += booked_days

    # Debugging: Check total nights after adjustment
    print(f"Total nights after adjustment (including checkout day): {total_nights}")

    # Calculate ADR
    adr = round(total_earnings / total_nights, 2) if total_nights > 0 else 0

    # Debugging: Final ADR calculation
    print(f"ADR for house {house_id} in {month}/{year}: {adr}")

    return adr




def calculate_booking_trends(house_id, year):
    if not house_id:
        return []

    monthly_bookings = Booking.objects.filter(house_id=house_id, start_date__year=year) \
        .annotate(month=F('start_date__month')) \
        .values('month') \
        .annotate(total=Count('id')) \
        .order_by('month')

    trends = {i: 0 for i in range(1, 13)}
    for entry in monthly_bookings:
        trends[entry['month']] = entry['total']

    return [trends[i] for i in range(1, 13)]


from datetime import datetime, timedelta
from django.db.models import Count, F
import calendar

# Add the necessary calculations in the view
def calculate_longest_booking(house_id, month, year):
    # Filter by month and year always
    filters = {
        "start_date__year": year,
        "start_date__month": month
    }

    # Only filter by house if a specific house_id is given
    if house_id is not None:
        filters["house_id"] = house_id

    bookings = Booking.objects.filter(**filters)

    longest_booking = 0
    for booking in bookings:
        booking_length = (booking.end_date - booking.start_date).days + 1
        longest_booking = max(longest_booking, booking_length)

    return longest_booking

def calculate_average_booking_length(house_id, month, year):
    filters = {
        "start_date__year": year,
        "start_date__month": month
    }
    if house_id is not None:
        filters["house_id"] = house_id

    bookings = Booking.objects.filter(**filters)
    total_days = 0
    total_bookings = bookings.count()
    for booking in bookings:
        total_days += (booking.end_date - booking.start_date).days + 1
    return round(total_days / total_bookings, 2) if total_bookings else 0

def calculate_booking_length_distribution(house_id, month, year):
    filters = {
        "start_date__year": year,
        "start_date__month": month
    }
    if house_id is not None:
        filters["house_id"] = house_id

    bookings = Booking.objects.filter(**filters)
    distribution = {"1-3 days": 0, "4-7 days": 0, "8+ days": 0}
    for booking in bookings:
        booking_length = (booking.end_date - booking.start_date).days + 1
        if booking_length <= 3:
            distribution["1-3 days"] += 1
        elif 4 <= booking_length <= 7:
            distribution["4-7 days"] += 1
        else:
            distribution["8+ days"] += 1
    total_bookings = bookings.count()
    distribution_percent = {k: (v / total_bookings) * 100 if total_bookings else 0 for k, v in distribution.items()}
    return distribution_percent



def calculate_bookings_this_week(house_id):
    today = datetime.today()
    start_of_week = today - timedelta(days=today.weekday())  # Monday of this week
    end_of_week = start_of_week + timedelta(days=6)  # Sunday of this week
    return Booking.objects.filter(house_id=house_id, start_date__gte=start_of_week, start_date__lte=end_of_week).count()

def calculate_bookings_last_week(house_id):
    today = datetime.today()
    start_of_last_week = today - timedelta(days=today.weekday() + 7)  # Monday of last week
    end_of_last_week = start_of_last_week + timedelta(days=6)  # Sunday of last week
    return Booking.objects.filter(house_id=house_id, start_date__gte=start_of_last_week, start_date__lte=end_of_last_week).count()


def calculate_most_common_booking_days(house_id, month, year):
    filters = {
        "start_date__year": year,
        "start_date__month": month
    }
    if house_id is not None:
        filters["house_id"] = house_id

    bookings = Booking.objects.filter(**filters)
    day_counts = {i: 0 for i in range(7)}  # 0 = Monday, ..., 6 = Sunday

    for booking in bookings:
        day_of_week = booking.start_date.weekday()
        day_counts[day_of_week] += 1

    if all(count == 0 for count in day_counts.values()):
        return "No bookings"

    most_common_day = max(day_counts, key=day_counts.get)
    return calendar.day_name[most_common_day]


from datetime import datetime
import calendar
from .models import House  # Adjust the import as per your project structure
from datetime import datetime
from .models import Booking, House
import calendar

def house_compare(request):
    current_date = datetime.now()
    current_year = current_date.year

    selected_month = int(request.GET.get('month', current_date.month))  # Default to current month
    selected_year = int(request.GET.get('year', current_year))  # Default to current year
    #house_id = request.GET.get('house', '')  # Get house_id from the URL query parameters
    house_id_str = request.GET.get('house', '')
    house_id = int(house_id_str) if house_id_str.isdigit() else None

    if selected_month == 0:
        selected_month = current_date.month  # Set to the current month if "all months" is selected

    houses = House.objects.all()  # Get all houses for the table/graphs

    house_profit_data = []
    booking_trends = [0] * 12  # Initialize an empty trend array

    selected_house_data = None  # For storing specific metrics if a house is selected

    # Calculate total bookings for the selected month and year
    total_bookings_per_house = {}  # Dictionary to store the total bookings per house

    for house in houses:
        # Calculate profit data and other metrics for all houses
        profit_data = calculate_profit(house.id, selected_month, selected_year)
        occupancy_rate = calculate_occupancy_rate(house.id, selected_month, selected_year)
        adr = calculate_adr(house.id, selected_month, selected_year)
        house_trends = calculate_booking_trends(house.id, selected_year)

        # Calculate total bookings for the selected month and year for the house
        total_bookings = Booking.objects.filter(
            house=house,
            start_date__month=selected_month,
            start_date__year=selected_year
        ).count()

        total_bookings_per_house[house.id] = total_bookings  # Store total bookings for this house

        # Accumulate bookings across houses
        booking_trends = [booking_trends[i] + house_trends[i] for i in range(12)]

        # Store the data for all houses in the table/graphs
        house_profit_data.append({
            "house": house,
            "total_earnings": profit_data["total_earnings"],
            "total_expenses": profit_data["total_expenses"],
            "profit": profit_data["profit"],
            "occupancy_rate": occupancy_rate,
            "adr": adr,
            "total_bookings": total_bookings,  # Add total bookings for the house
            "booking_trends": house_trends,  # Store per-house trends
        })

        # If a specific house is selected, gather the detailed metrics
        if house_id is not None and house.id == house_id:
            longest_booking = calculate_longest_booking(house.id, selected_month, selected_year)
            average_booking_length = calculate_average_booking_length(house.id, selected_month, selected_year)
            booking_length_distribution = calculate_booking_length_distribution(house.id, selected_month, selected_year)
            most_common_booking_day = calculate_most_common_booking_days(house.id, selected_month, selected_year)

            # Store the selected house data for metrics display
            selected_house_data = {
                "house": house,
                "longest_booking": longest_booking,
                "average_booking_length": average_booking_length,
                "booking_length_distribution": booking_length_distribution,
                "total_bookings": total_bookings,  # Display total bookings for the selected house
                "most_common_booking_day": most_common_booking_day
            }

    # If no house is selected, show a message
    if not house_id:
        # Aggregate metrics for all houses
        longest_booking = calculate_longest_booking(None, selected_month, selected_year)
        average_booking_length = calculate_average_booking_length(None, selected_month, selected_year)
        booking_length_distribution = calculate_booking_length_distribution(None, selected_month, selected_year)
        most_common_booking_day = calculate_most_common_booking_days(None, selected_month, selected_year)

        total_bookings = Booking.objects.filter(
            start_date__month=selected_month,
            start_date__year=selected_year
        ).count()

        selected_house_data = {
            "house": {"name": "All Houses"},
            "longest_booking": longest_booking,
            "average_booking_length": average_booking_length,
            "booking_length_distribution": booking_length_distribution,
            "total_bookings": total_bookings,
            "most_common_booking_day": most_common_booking_day
    }


    return render(request, "houses/house_compare.html", {
        "houses": houses,
        "house_profit_data": house_profit_data,  # Table/graphs for all houses
        "selected_house_data": selected_house_data,  # Metrics for the selected house or message
        "selected_month": selected_month,
        "selected_year": selected_year,
        "months": [{"value": i, "name": calendar.month_name[i]} for i in range(1, 13)],
        "years": range(current_year, current_year + 5),
        "booking_trends": booking_trends,  # Accumulated trend for all houses
        "total_bookings_per_house": total_bookings_per_house,  # Pass total bookings to the template
    })



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from .models import House, Discount
from .forms import DiscountForm  
from django.shortcuts import render, redirect
from .models import House, Discount
from .forms import DiscountForm  # Ensure this form is correctly defined

from django.contrib import messages

def set_discount(request):
    if request.method == 'POST':
        form = DiscountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Discount successfully set!")
            return redirect('discounts_page')
        else:
            messages.error(request, "There was an error in the form.")
    else:
        form = DiscountForm()

    return render(request, 'houses/discounts.html', {'form': form})

def discounts_page(request):
    """Display discounts and allow setting new ones."""
    houses = House.objects.all()
    discounts = Discount.objects.all()  # Fetch existing discounts

    if request.method == 'POST':
        form = DiscountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Discount successfully set!")
            return redirect('discounts_page')
        else:
            messages.error(request, "There was an error in the form.")
    else:
        form = DiscountForm()

    return render(request, 'houses/discounts.html', {
        'houses': houses,
        'discounts': discounts,  # Pass discounts to template
        'form': form,
    })

def edit_discount(request, discount_id):
    """Edit an existing discount."""
    discount = get_object_or_404(Discount, id=discount_id)
    
    if request.method == 'POST':
        form = DiscountForm(request.POST, instance=discount)
        if form.is_valid():
            form.save()
            messages.success(request, "Discount successfully updated!")
            return redirect('discounts_page')
        else:
            messages.error(request, "There was an error updating the discount.")
    else:
        form = DiscountForm(instance=discount)

    return render(request, 'houses/edit_discount.html', {'form': form, 'discount': discount})

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Discount

def delete_discount(request, discount_id):
    """Delete a discount using AJAX."""
    if request.method == 'DELETE':
        discount = get_object_or_404(Discount, id=discount_id)
        discount.delete()
        return JsonResponse({'message': 'Discount deleted successfully!'})
    return JsonResponse({'error': 'Invalid request'}, status=400)


# views.py
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from datetime import datetime
from django.utils import timezone
from .models import House, Discount

from datetime import datetime, timedelta
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import House, Discount

def get_discounted_price(request):
    house_id = request.GET.get('house_id')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    # Get the house object
    house = get_object_or_404(House, id=house_id)

    # Get applicable discounts for the house within the given date range
    applicable_discounts = Discount.objects.filter(
        house=house,
        start_date__lte=end_date,
        end_date__gte=start_date
    )

    total_discounted_price = 0
    total_non_discounted_price = 0
    current_date = start_date

    while current_date <= end_date:
        # Find if this day has a discount
        daily_discount = False
        for discount in applicable_discounts:
            if discount.start_date <= current_date <= discount.end_date:
                # Apply discount for this day
                discount_amount = (discount.discount_percentage / 100) * house.price
                daily_discount = True
                total_discounted_price += (house.price - discount_amount)
                break
        
        if not daily_discount:
            # No discount for this day, apply normal price
            total_non_discounted_price += house.price
        
        # Move to the next day
        current_date += timedelta(days=1)

    # Total price is the sum of discounted and non-discounted days
    total_price = total_discounted_price + total_non_discounted_price

    # Debug logs for better visibility
    print(f"House ID: {house_id}")
    print(f"Start Date: {start_date_str}")
    print(f"End Date: {end_date_str}")
    print(f"Total Discounted Price: {total_discounted_price}")
    print(f"Total Non-Discounted Price: {total_non_discounted_price}")
    print(f"Total Price: {total_price}")

    # Format the price to 2 decimals
    total_price = round(total_price, 2)
    total_discounted_price = round(total_discounted_price, 2)
    total_non_discounted_price = round(total_non_discounted_price, 2)

    return JsonResponse({
        'total_price': total_price,
        'total_discounted_price': total_discounted_price,
        'total_non_discounted_price': total_non_discounted_price
    })



from django.http import JsonResponse
from .models import House, Discount
from django.http import JsonResponse
from .models import House, Discount
from datetime import datetime

def get_discounted_price_for_day(request):
    house_id = request.GET.get('house_id')
    date_str = request.GET.get('date')  # Expecting date in 'YYYY-MM-DD' format

    # Validate inputs
    if not house_id or not date_str:
        return JsonResponse({'error': 'Missing house_id or date parameters'}, status=400)

    try:
        house = House.objects.get(id=house_id)
    except House.DoesNotExist:
        return JsonResponse({'error': 'House not found'}, status=404)

    try:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()  # Convert string to date object
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    # Default price is the house's regular price
    discounted_price = house.price

    # Check for a discount that applies to the selected date
    discount = Discount.objects.filter(house=house, start_date__lte=date, end_date__gte=date).first()
    if discount:
        # Calculate the discounted price
        discount_amount = (discount.discount_percentage / 100) * house.price
        discounted_price = house.price - discount_amount  # Apply the discount

    # Make sure discounted_price is a valid number before returning it
    try:
        discounted_price = float(discounted_price)  # Ensure it's a float
    except ValueError:
        return JsonResponse({'error': 'Invalid price value'}, status=500)

    # Return the discounted price as JSON
    return JsonResponse({'discounted_price': round(discounted_price, 2)})  # Return rounded to 2 decimal places


from django.shortcuts import render, redirect
from .models import CleaningFeeSetting
from .forms import CleaningFeeForm



def edit_cleaning_fee(request):
    setting, _ = CleaningFeeSetting.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        form = CleaningFeeForm(request.POST, instance=setting)
        if form.is_valid():
            form.save()
            # după salvare afișăm modalul de succes
            return render(request, 'houses/edit_cleaning_fee.html', {'form': form, 'success': True})
    else:
        form = CleaningFeeForm(instance=setting)
        
    return render(request, 'houses/edit_cleaning_fee.html', {'form': form})
